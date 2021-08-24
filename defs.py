import win32com.client as win32
import global_var as gvar
import openpyxl
import time

def window_change(DataWindow, CloseWindow):  # 複製catia檔案(留著的視窗,被複製的視窗)
    time.sleep(1)
    catapp = win32.Dispatch("CATIA.Application")
    partdoc = catapp.ActiveDocument
    selection1 = partdoc.Selection
    part1 = partdoc.Part
    relations1 = part1.Relations
    formulal_Count = part1.Relations.Count
    selection1.Clear()
    for form_number in range(1, formulal_Count + 1):
        formula1 = relations1.Item(form_number)
        print(formula1)
        selection1.Add(formula1)
        time.sleep(0.1)
    parameters2 = part1.Parameters
    parameter_count = parameters2.RootParameterSet.DirectParameters.Count
    for parame_number in range(1, parameter_count + 1):
        paramet1 = parameters2.RootParameterSet.DirectParameters.Item(parame_number)
        selection1.Add(paramet1)
        time.sleep(0.1)
    bodies1 = part1.Bodies
    bodies_Count = part1.Bodies.Count
    for bodies_number in range(1, bodies_Count + 1):
        bodie1 = bodies1.Item(bodies_number)
        selection1.Add(bodie1)
        time.sleep(0.1)
    AxisSystems1 = part1.AxisSystems
    AxisSystems_Count = part1.AxisSystems.Count
    for Axis_number in range(1, AxisSystems_Count + 1):
        Axis1 = AxisSystems1.Item(Axis_number)
        selection1.Add(Axis1)
        time.sleep(0.1)
    hybridBody_Count = part1.HybridBodies.Count
    for hybridBody_number in range(1, hybridBody_Count + 1):
        hybridBody1 = part1.HybridBodies.Item(hybridBody_number)
        selection1.Add(hybridBody1)
        time.sleep(0.1)
    time.sleep(1)
    selection1.Copy()
    time.sleep(1)
    window = catapp.Windows
    # PasteWindow = window.Item(DataWindow)
    DataWindow.Activate()
    catapp = win32.Dispatch("CATIA.Application")
    partdoc = catapp.ActiveDocument
    selection2 = partdoc.Selection
    part1 = partdoc.part
    selection2.Clear()
    selection2.Add(part1)
    selection2.Paste()
    time.sleep(1)
    selection2.Clear()
    # CloseWin = window.Item(CloseWindow)
    CloseWindow.Activate()
    partdoc = catapp.ActiveDocument
    partdoc.Close()
    time.sleep(1)


def part_open(input_root, dir):
    # 連結CATIA
    catapp = win32.Dispatch("CATIA.Application")
    document = catapp.Documents
    # 將路徑設為目錄的文字宣告
    # gvar.folderdir = directory
    # 定義零件檔檔名
    part_dir = input_root + dir
    print(part_dir)
    # partdoc = document.Open("%s%s.%s" % (directory,target,"CATPart"))
    # 開啟該零件檔
    partdoc = document.Open(part_dir)


def environment_set(PartFileName, BodyName, HybridBodyName):
    catapp = win32.Dispatch("CATIA.Application")
    productDocument1 = catapp.ActiveDocument
    ElementProduct = productDocument1
    documents1 = catapp.Documents
    partDocument1 = documents1.Item(PartFileName + ".CATPart")
    ElementDocument = partDocument1
    selection1 = ElementDocument.Selection
    part1 = partDocument1.Part
    bodies1 = part1.Bodies
    body1 = bodies1.Item(BodyName)
    ElementBody = body1
    sketches1 = body1.Sketches
    hybridBodies1 = part1.HybridBodies
    hybridBody1 = hybridBodies1.Item(HybridBodyName)
    ElementHybridBody = hybridBody1
    return ElementProduct, ElementDocument, ElementBody, ElementHybridBody


def BuildXYZpoint(X, Y, Z, PointName, ElementDocument, ElementHybridBody):
    part1 = ElementDocument.Part
    hybridShapeFactory1 = part1.HybridShapeFactory
    hybridShapePointCoord1 = hybridShapeFactory1.AddNewPointCoord(5, 10, 0)
    ElementHybridBody.AppendHybridShape(hybridShapePointCoord1)
    part1.InWorkObject = hybridShapePointCoord1
    hybridShapePointCoord1.X.Value = X
    hybridShapePointCoord1.Y.Value = Y
    hybridShapePointCoord1.Z.Value = Z
    part1.Update()
    ElementPoint = hybridShapePointCoord1
    hide(hybridShapePointCoord1, ElementDocument)
    hybridShapePointCoord1.Name = PointName
    return ElementPoint


def hide(HideElement, ElementDocument):
    selection1 = ElementDocument.Selection
    visPropertySet1 = selection1.VisProperties
    selection1.Clear()
    selection1.Add(HideElement)
    time.sleep(1)
    visPropertySet1 = visPropertySet1.Parent
    visPropertySet1.SetShow(1)
    selection1.Clear()


def BuildPointChose(point_type, E_Reference, ElementDocument, ElementHybridBody, SketchPosition, ElementBody):
    part1 = ElementDocument.part
    hybridShapeFactory1 = part1.HybridShapeFactory
    hybridShapes1 = ElementHybridBody.HybridShapes
    reference1 = part1.CreateReferenceFromObject(E_Reference)
    if point_type == "Center_Cruve":
        hybridShapePointCenter1 = hybridShapeFactory1.AddNewPointCenter(E_Reference)
        out_put_element = hybridShapePointCenter1
    elif point_type == "Cruve_ratio":
        hybridShapePointOnCurve1 = hybridShapeFactory1.AddNewPointOnCurveFromPercent(reference1, 0, True)
        out_put_element = hybridShapePointOnCurve1
    else:
        raise ValueError('Incorrect Target Type')
    if SketchPosition == "Body":
        ElementBody.InsertHybridShape(out_put_element)
    elif SketchPosition == "Hybridbody":
        ElementHybridBody.AppendHybridShape(out_put_element)
    ElementPoint5 = out_put_element
    part1.InWorkObject = ElementPoint5
    return ElementPoint5


def JoinElement(Reference1, Reference2, ElementDocument, ElementBody, ElementHybridBody,
                SketchPosition):  # '(元素1,元素2) out=element_Reference(1)
    time.sleep(0.1)
    part1 = ElementDocument.part
    hybridShapeFactory1 = part1.HybridShapeFactory
    reference1 = part1.CreateReferenceFromObject(Reference1)
    reference2 = part1.CreateReferenceFromObject(Reference2)
    hybridShapeAssemble1 = hybridShapeFactory1.AddNewJoin(reference1, reference2)
    hybridShapeAssemble1.SetConnex(1)
    hybridShapeAssemble1.SetManifold(1)
    hybridShapeAssemble1.SetSimplify(0)
    hybridShapeAssemble1.SetSuppressMode(0)
    hybridShapeAssemble1.SetDeviation(0.001)
    hybridShapeAssemble1.SetAngularToleranceMode(0)
    hybridShapeAssemble1.SetAngularTolerance(0.5)
    hybridShapeAssemble1.SetFederationPropagation(0)
    if SketchPosition == "Body":
        ElementBody.InsertHybridShape(hybridShapeAssemble1)
    if SketchPosition == "Hybridbody":
        ElementHybridBody.AppendHybridShape(hybridShapeAssemble1)
    ElementReference = hybridShapeAssemble1
    part1.InWorkObject = hybridShapeAssemble1
    part1.Update()
    return ElementReference


def break_relationship(BreakElement, ElementType, ElementDocument, ElementBody, ElementHybridBody,
                       SketchPosition):  # (需打斷關係之元素,element_type="point" or "line")   element_Reference(5)為out
    part1 = ElementDocument.Part
    hybridShapeFactory1 = part1.HybridShapeFactory
    reference5 = part1.CreateReferenceFromObject(BreakElement)
    if ElementType == 'line':
        hybridShape_break = hybridShapeFactory1.AddNewLineDatum(reference5)
    elif ElementType == 'point':
        hybridShape_break = hybridShapeFactory1.AddNewPointDatum(reference5)
    else:
        hybridShape_break = hybridShapeFactory1.AddNewCurveDatum(reference5)
    if SketchPosition == ' Body':
        ElementBody.InsertHybridShape(hybridShape_break)
    elif SketchPosition == 'Hybridbody':
        ElementHybridBody.AppendHybridShape(hybridShape_break)
    else:
        raise ValueError('Incorrect Target Type')

    part1.InWorkObject = hybridShape_break
    part1.Update()
    ElementReference5 = hybridShape_break
    return ElementReference5


def delete_object(delete_element, ElementProduct):
    selection1 = ElementProduct.Selection
    selection1.Clear()
    selection1.Add(delete_element)
    if selection1.Count != 0:
        selection1.Delete()
        selection1.Clear()


def BuildSketch(SketchName, PlaneElement, ElementDocument, SketchPosition, ElementBody,
                ElementHybridBody):  # (sketch名稱,產生sketch的平面)  element_sketch(1)為產生出來的草圖
    part1 = ElementDocument.Part
    reference1 = part1.CreateReferenceFromObject(PlaneElement)
    if SketchPosition == "Body":
        sketch1 = ElementBody.Sketches.Add(reference1)
    elif SketchPosition == 'Hybridbody':
        sketch1 = ElementHybridBody.HybridSketches.Add(reference1)
    else:
        raise ValueError('Incorrect Target Type')
    geometricElements1 = sketch1.GeometricElements
    axis2D1 = geometricElements1.Item("AbsoluteAxis")
    part1.InWorkObject = sketch1
    factory2D1 = sketch1.OpenEdition()
    line2D1 = axis2D1.getItem("HDirection")
    line2D2 = axis2D1.getItem("VDirection")
    sketch1.CloseEdition()
    sketch1.Name = SketchName
    ElementSketch = sketch1
    return ElementSketch


def SketchBuildCallout(MainSketch, Direction, CalloutType, Data, ElementDocument, ElementPoint, NextPoint):
    part1 = ElementDocument.Part
    factory2D1 = MainSketch.OpenEdition()
    geometricElements1 = MainSketch.GeometricElements
    axis2D1 = geometricElements1.Item("AbsoluteAxis")
    line2D1 = axis2D1.getItem("HDirection")
    line2D2 = axis2D1.getItem("VDirection")
    reference1 = part1.CreateReferenceFromObject(ElementPoint)
    if Direction != 'Radius':
        try:
            reference2 = part1.CreateReferenceFromObject(NextPoint)
        except:
            MainSketch.CloseEdition()
            reference2 = part1.CreateReferenceFromObject(NextPoint)
            factory2D1 = MainSketch.OpenEdition()
    reference3 = part1.CreateReferenceFromObject(line2D1)  # 水平方向
    reference4 = part1.CreateReferenceFromObject(line2D2)  # 垂直方向
    constraints1 = MainSketch.Constraints
    if Direction == 'Horizontal':
        constraint1 = constraints1.AddTriEltCst(1, reference1, reference2, reference3)
    elif Direction == "Vertical":
        constraint1 = constraints1.AddTriEltCst(1, reference1, reference2, reference4)
    elif Direction == 'free':
        constraint1 = constraints1.AddBiEltCst(1, reference1, reference2)
    elif Direction == "Radius":
        constraint1 = constraints1.AddMonoEltCst(14, reference1)
    else:
        raise ValueError('Incorrect Target Type')

    if CalloutType == "Callout":
        constraint1.mode = 1
        Data = constraint1.dimension.Value
    elif CalloutType == 'Binding':
        constraint1.mode = 0
        constraint1.dimension.Value = Data
    MainSketch.CloseEdition()
    return Data


def SketchHidePoint(MainSketch, CirclcPoint, X, Y, Construct, ElementDocument, ElementSketch):
    part1 = ElementDocument.part
    part1.InWorkObject = MainSketch
    factory2D1 = MainSketch.OpenEdition()
    geometricElements1 = MainSketch.GeometricElements
    axis2D1 = geometricElements1.Item("AbsoluteAxis")
    line2D5 = axis2D1.getItem("HDirection")
    line2D6 = axis2D1.getItem("VDirection")
    point2D1 = factory2D1.CreatePoint(10, 10)
    point2D1.ReportName = 1
    if Construct == 'True':
        point2D1.Construction = True
    elif Construct == '"False"':
        point2D1.Construction = False
    ElementPoint = point2D1
    NextPoint = CirclcPoint
    SketchBuildCallout(ElementSketch, "Horizontal", "Binding", X, ElementDocument, ElementPoint, NextPoint)
    SketchBuildCallout(ElementSketch, "Vertical", "Binding", Y, ElementDocument, ElementPoint, NextPoint)
    ElementPoint30 = point2D1
    MainSketch.CloseEdition()
    part1.InWorkObject = MainSketch
    return ElementPoint30


def SketchCircle(MainSketch, CirclcPoint, R, ElementDocument):
    part1 = ElementDocument.Part
    part1.InWorkObject = MainSketch
    factory2D1 = MainSketch.OpenEdition()
    geometricElements1 = MainSketch.GeometricElements
    axis2D1 = geometricElements1.Item("AbsoluteAxis")
    line2D5 = axis2D1.getItem("HDirection")
    line2D6 = axis2D1.getItem("VDirection")
    circle2D1 = factory2D1.CreateClosedCircle(-552.102722, 144.968857, 10)
    circle2D1.CenterPoint = CirclcPoint
    constraints1 = MainSketch.Constraints
    reference2 = part1.CreateReferenceFromObject(circle2D1)
    constraint1 = constraints1.AddMonoEltCst(14, reference2)
    constraint1.mode = 0
    length1 = constraint1.dimension
    length1.Value = R
    ElementReference11 = circle2D1
    MainSketch.CloseEdition()
    part1.InWorkObject = MainSketch
    return ElementReference11


def ProjectionLine(ElementReference11, ElementReference12, ElementDocument, ElementHybridBody, ElementBody,
                   SketchPosition, Display):
    part1 = ElementDocument.Part
    hybridShapes1 = ElementHybridBody.HybridShapes
    hybridShapeFactory1 = part1.HybridShapeFactory  # 'line
    reference1 = part1.CreateReferenceFromObject(ElementReference11)  # 'plane
    reference2 = part1.CreateReferenceFromObject(ElementReference12)
    hybridShapeProject1 = hybridShapeFactory1.AddNewProject(reference1, reference2)
    hybridShapeProject1.SolutionType = 0
    hybridShapeProject1.Normal = True
    hybridShapeProject1.SmoothingType = 0
    ElementBody.InsertHybridShape(hybridShapeProject1)
    part1.InWorkObject = hybridShapeProject1
    part1.Update()
    reference3 = part1.CreateReferenceFromObject(hybridShapeProject1)
    hybridShapeCurveExplicit1 = hybridShapeFactory1.AddNewCurveDatum(reference3)
    out_put_element = hybridShapeCurveExplicit1
    if SketchPosition == "Body":
        ElementBody.InsertHybridShape(out_put_element)
    elif SketchPosition == "Hybridbody":
        ElementHybridBody.AppendHybridShape(out_put_element)
    ElementLine5 = out_put_element
    part1.InWorkObject = out_put_element
    part1.Update()
    hybridShapeFactory1.DeleteObjectForDatum(reference3)
    if Display == 'True':
        pass
    elif Display == 'False':
        selection1 = ElementDocument.Selection
        visPropertySet1 = selection1.VisProperties
        selection1.Clear()
        selection1.add(ElementLine5)
        visPropertySet1 = visPropertySet1.Parent
        visPropertySet1.SetShow(1)
        selection1.Clear()
    return ElementLine5


def ElementExtremumFourPoint(element, ElementDocument, ElementBody, ElementHybridBody):
    (ElementReference1) = ExtremumPoint("X_min", "Y_min", "Z_max", 2, element, ElementDocument, ElementBody,
                                        ElementHybridBody)
    ElementPoint21 = ElementReference1
    ElementPoint21.Name = str(element.Name) + "_Xmin"
    (ElementReference1) = ExtremumPoint("X_max", "Y_min", "Z_max", 2, element, ElementDocument, ElementBody,
                                        ElementHybridBody)
    ElementPoint22 = ElementReference1
    ElementPoint22.Name = str(element.Name) + "_Xmax"
    (ElementReference1) = ExtremumPoint("Y_min", "X_min", "Z_max", 2, element, ElementDocument, ElementBody,
                                        ElementHybridBody)
    ElementPoint23 = ElementReference1
    ElementPoint23.Name = str(element.Name) + "_Ymin"
    (ElementReference1) = ExtremumPoint("Y_max", "X_min", "Z_max", 2, element, ElementDocument, ElementBody,
                                        ElementHybridBody)
    ElementPoint24 = ElementReference1
    ElementPoint24.Name = str(element.Name) + "_Ymax"
    return ElementPoint21, ElementPoint22, ElementPoint23, ElementPoint24


def ExtremumPoint(Direction1, Direction2, Direction3, DirectionNumber, element, ElementDocument, ElementBody,
                  ElementHybridBody):
    part1 = ElementDocument.Part
    hybridShapes2 = ElementBody.HybridShapes
    hybridShapeFactory1 = part1.HybridShapeFactory
    X_min = hybridShapeFactory1.AddNewDirectionByCoord(1, 0, 0)
    X_max = hybridShapeFactory1.AddNewDirectionByCoord(-1, 0, 0)
    Y_min = hybridShapeFactory1.AddNewDirectionByCoord(0, 1, 0)
    Y_max = hybridShapeFactory1.AddNewDirectionByCoord(0, -1, 0)
    Z_max = hybridShapeFactory1.AddNewDirectionByCoord(0, 0, 1)
    Z_min = hybridShapeFactory1.AddNewDirectionByCoord(0, 0, -1)
    zero = hybridShapeFactory1.AddNewDirectionByCoord(0, 0, 0)
    if Direction1 == "X_min":
        direction_1 = X_min
    elif Direction1 == "X_max":
        direction_1 = X_max
    elif Direction1 == "Y_min":
        direction_1 = Y_min
    elif Direction1 == "Y_max":
        direction_1 = Y_max
    elif Direction1 == "Z_min":
        direction_1 = Z_min
    elif Direction1 == "Z_max":
        direction_1 = Z_max
    if Direction2 == "X_min":
        direction_2 = X_min
    elif Direction2 == "X_max":
        direction_2 = X_max
    elif Direction2 == "Y_min":
        direction_2 = Y_min
    elif Direction2 == "Y_max":
        direction_2 = Y_max
    elif Direction2 == "Z_min":
        direction_2 = Z_min
    elif Direction2 == "Z_max":
        direction_2 = Z_max
    if Direction3 == "X_min":
        direction_3 = X_min
    elif Direction3 == "X_max":
        direction_3 = X_max
    elif Direction3 == "Y_min":
        direction_3 = Y_min
    elif Direction3 == "Y_max":
        direction_3 = Y_max
    elif Direction3 == "Z_min":
        direction_3 = Z_min
    elif Direction3 == "Z_max":
        direction_3 = Z_max
    reference2 = part1.CreateReferenceFromObject(element)
    hybridShapeExtremum1 = hybridShapeFactory1.AddNewExtremum(reference2, zero, 0)
    if DirectionNumber == 1:
        hybridShapeExtremum1.direction = direction_1
    elif DirectionNumber == 2:
        hybridShapeExtremum1.direction = direction_1
        hybridShapeExtremum1.Direction2 = direction_2
    elif DirectionNumber == 3:
        hybridShapeExtremum1.direction = direction_1
        hybridShapeExtremum1.Direction2 = direction_2
        hybridShapeExtremum1.Direction3 = direction_3
    else:
        raise ValueError('Incorrect Target Type')
    hybridShapeExtremum1.ExtremumType2 = 0
    ElementHybridBody.AppendHybridShape(hybridShapeExtremum1)
    ElementReference1 = hybridShapeExtremum1
    return ElementReference1


def BuildPoint(ElementReference, ElementDocument, ElementBody, ElementHybridBody, SketchPosition):
    part1 = ElementDocument.Part
    hybridShapes1 = ElementHybridBody.HybridShapes
    hybridShapeFactory1 = part1.HybridShapeFactory
    hybridShapePointCoord1 = hybridShapeFactory1.AddNewPointCoord(0, 0, 0)
    reference1 = part1.CreateReferenceFromObject(ElementReference)
    hybridShapePointCoord1.PtRef = reference1
    if SketchPosition == "Body":
        ElementBody.InsertHybridShape(hybridShapePointCoord1)
    elif SketchPosition == "Hybridbody":
        ElementHybridBody.AppendHybridShape(hybridShapePointCoord1)
    part1.InWorkObject = hybridShapePointCoord1
    part1.Update()
    selection1 = ElementDocument.Selection
    visPropertySet1 = selection1.VisProperties
    selection1.Clear()
    selection1.add(hybridShapePointCoord1)
    visPropertySet1 = visPropertySet1.Parent
    visPropertySet1.SetShow(1)
    selection1.Clear()
    return hybridShapePointCoord1


def AngleMeasure(centor_point_element, measure_angle_point, ElementDocument, ElementHybridBody):
    part1 = ElementDocument.Part
    relations1 = part1.Relations  # 關聯指令起手宣告
    parameters1 = part1.Parameters  # 參數指令起手宣告
    hybridShapeFactory1 = part1.HybridShapeFactory  # 曲面指令起手宣告
    hybridShapes1 = ElementHybridBody.HybridShapes  # 依據指令起手宣告
    selection1 = ElementDocument.Selection  # 選擇宣告
    # -----------------------------------------------------------------------------起手是宣告
    selection1.Clear()
    selection1.Search("Name=measure_angle_P_*,all")
    # -----------------------------------------------------------------------------建立測量參數
    if    selection1.Count != 2    :
        length1 = parameters1.CreateDimension("", "ANGLE", 0 )    #build parameter
        length1.rename( "measure_angle_P_Parameter")
        formula1 = relations1.CreateFormula("measure_formula", "", length1, "length( )")
        formula1.rename(
        "measure_angle_P_Formula")
    else:
        length1 = parameters1.Item("measure_angle_P_Parameter")
        formula1 = relations1.Item("measure_angle_P_Formula")
    # -----------------------------------------------------------------------------建立測量參數
    # -----------------------------------------------------------------------------建立測量依據點
    selection1.Clear()
    selection1.Search("Name=measure_angle_*_point,all")
    if    selection1.Count != 2:
        reference1 = part1.CreateReferenceFromObject(centor_point_element)
        (element_point5)=BuildXYZpoint(50, 0, 0, "measure_angle_H_point",ElementDocument,ElementHybridBody)  # (X座標,Y座標,Z座標) element_point(5) 為out
        element_point5.PtRef = reference1
        element_point5.Name = "measure_angle_H_point"
        H_point_element = element_point5
        (element_point5)=BuildXYZpoint(0, 50, 0, "measure_angle_V_point",ElementDocument,ElementHybridBody)  # (X座標,Y座標,Z座標) element_point(5) 為out
        element_point5.PtRef = reference1
        element_point5.Name = "measure_angle_V_point"
        V_point_element = element_point5
    else:
        H_point_element = hybridShapes1.Item("measure_angle_H_point")
        V_point_element = hybridShapes1.Item("measure_angle_V_point")
    # -----------------------------------------------------------------------------建立測量依據點
    formula1.Modify(
    "angle(`" + ElementHybridBody.Name + "\\" + centor_point_element.Name + "`, `" + ElementHybridBody.Name + "\\" + H_point_element.Name + "`, `" + ElementHybridBody.Name + "\\" + measure_angle_point.Name + "`) ")
    part1.UpdateObject(formula1)
    angle_1_now = length1.Value
    formula1.Modify(
    "angle(`" + ElementHybridBody.Name + "\\" + centor_point_element.Name + "`, `" + ElementHybridBody.Name + "\\" + V_point_element.Name + "`, `" + ElementHybridBody.Name + "\\" + measure_angle_point.Name + "`) ")
    part1.UpdateObject(formula1)
    angle_2_now = length1.Value
    # judgement_quadrant 判斷該元素在第幾象限(水平判斷角,垂直判斷角,輸出第幾象限)
    if angle_1_now > 90 and angle_2_now > 90 :
        quadrant_number = 3
    elif angle_1_now > 90 and angle_2_now < 90:
        quadrant_number = 2
    elif angle_1_now < 90 and angle_2_now > 90:
        quadrant_number = 4
    elif angle_1_now < 90 and angle_2_now < 90:
        quadrant_number = 1
    # judgement_quadrant 判斷該元素在第幾象限(水平判斷角,垂直判斷角,輸出第幾象限)
    if    angle_1_now > 90:
        angle_1_now = angle_1_now - 90
    if    quadrant_number == 3 or    quadrant_number == 4:
        angle_1_now = 90 - angle_1_now
    angel_out = angle_1_now + (quadrant_number - 1) * 90
    return angel_out


def tryangle(AngleNumber, ElementDocument, ElementSketch, CircleLineType, LineType, ElementHybridBody):
    part1 = ElementDocument.Part
    factory2D1 = ElementSketch.OpenEdition()
    geometricElements1 = ElementSketch.GeometricElements
    axis2D1 = geometricElements1.Item("AbsoluteAxis")
    line2D1 = axis2D1.getItem("HDirection")
    line2D1.ReportName = 1
    line2D2 = axis2D1.getItem("VDirection")
    line2D2.ReportName = 2
    point2D1 = factory2D1.CreatePoint(311.099756, 25.077236)
    point2D1.ReportName = 3
    point2D2 = factory2D1.CreatePoint(310.75246, 23.107621)
    point2D2.ReportName = 4
    line2D3 = factory2D1.CreateLine(311.099756, 25.077236, 310.75246, 23.107621)
    line2D3.ReportName = 5
    line2D3.StartPoint = point2D1
    line2D3.EndPoint = point2D2
    point2D3 = factory2D1.CreatePoint(308.782844, 23.454917)
    point2D3.ReportName = 6
    line2D4 = factory2D1.CreateLine(310.75246, 23.107621, 308.782844, 23.454917)
    line2D4.ReportName = 7
    line2D4.StartPoint = point2D2
    line2D4.EndPoint = point2D3
    point2D4 = factory2D1.CreatePoint(307.914603, 18.530878)
    point2D4.ReportName = 8
    line2D5 = factory2D1.CreateLine(308.782844, 23.454917, 307.914603, 18.530878)
    line2D5.ReportName = 9
    line2D5.StartPoint = point2D3
    line2D5.EndPoint = point2D4
    point2D5 = factory2D1.CreatePoint(325.015544, 15.515521)
    point2D5.ReportName = 10
    line2D6 = factory2D1.CreateLine(307.914603, 18.530878, 325.015544, 15.515521)
    line2D6.ReportName = 11
    line2D6.StartPoint = point2D4
    line2D6.EndPoint = point2D5
    point2D6 = factory2D1.CreatePoint(325.883784, 20.43956)
    point2D6.ReportName = 12
    line2D7 = factory2D1.CreateLine(325.015544, 15.515521, 325.883784, 20.43956)
    line2D7.ReportName = 13
    line2D7.StartPoint = point2D5
    line2D7.EndPoint = point2D6
    constraints1 = ElementSketch.Constraints
    reference2 = part1.CreateReferenceFromObject(point2D6)
    reference3 = part1.CreateReferenceFromObject(line2D4)
    constraint1 = constraints1.AddBiEltCst(2, reference2, reference3)
    constraint1.mode = 0
    point2D7 = factory2D1.CreatePoint(323.914169, 20.786856)
    point2D7.ReportName = 14
    line2D8 = factory2D1.CreateLine(325.883784, 20.43956, 323.914169, 20.786856)
    line2D8.ReportName = 15
    line2D8.StartPoint = point2D6
    line2D8.EndPoint = point2D7
    point2D8 = factory2D1.CreatePoint(324.261465, 22.756472)
    point2D8.ReportName = 16
    line2D9 = factory2D1.CreateLine(323.914169, 20.786856, 324.261465, 22.756472)
    line2D9.ReportName = 17
    line2D9.StartPoint = point2D7
    line2D9.EndPoint = point2D8
    point2D9 = factory2D1.CreatePoint(319.922297, 36.630088)
    point2D9.ReportName = 18
    point2D10 = factory2D1.CreatePoint(315.139627, 9.506219)
    point2D10.ReportName = 19
    line2D10 = factory2D1.CreateLine(319.922297, 36.630088, 315.139627, 9.506219)
    line2D10.ReportName = 20
    line2D10.Construction = True
    line2D10.StartPoint = point2D9
    line2D10.EndPoint = point2D10
    point2D11 = factory2D1.CreatePoint(309.157593, 33.844704)
    point2D11.ReportName = 21
    point2D12 = factory2D1.CreatePoint(343.239685, 33.844704)
    point2D12.ReportName = 22
    line2D11 = factory2D1.CreateLine(309.157593, 33.844704, 343.239685, 33.844704)
    line2D11.ReportName = 23
    line2D11.Construction = True
    line2D11.StartPoint = point2D11
    line2D11.EndPoint = point2D12
    reference4 = part1.CreateReferenceFromObject(line2D11)
    reference5 = part1.CreateReferenceFromObject(line2D1)
    constraint2 = constraints1.AddBiEltCst(10, reference4, reference5)
    constraint2.mode = 0
    reference6 = part1.CreateReferenceFromObject(line2D4)
    reference7 = part1.CreateReferenceFromObject(line2D3)
    constraint3 = constraints1.AddBiEltCst(11, reference6, reference7)
    constraint3.mode = 0
    reference8 = part1.CreateReferenceFromObject(line2D4)
    reference9 = part1.CreateReferenceFromObject(line2D5)
    constraint4 = constraints1.AddBiEltCst(11, reference8, reference9)
    constraint4.mode = 0
    reference10 = part1.CreateReferenceFromObject(line2D5)
    reference11 = part1.CreateReferenceFromObject(line2D6)
    constraint5 = constraints1.AddBiEltCst(11, reference10, reference11)
    constraint5.mode = 0
    reference12 = part1.CreateReferenceFromObject(line2D6)
    reference13 = part1.CreateReferenceFromObject(line2D7)
    constraint6 = constraints1.AddBiEltCst(11, reference12, reference13)
    constraint6.mode = 0
    reference14 = part1.CreateReferenceFromObject(line2D7)
    reference15 = part1.CreateReferenceFromObject(line2D8)
    constraint7 = constraints1.AddBiEltCst(11, reference14, reference15)
    constraint7.mode = 0
    reference16 = part1.CreateReferenceFromObject(line2D8)
    reference17 = part1.CreateReferenceFromObject(line2D9)
    constraint8 = constraints1.AddBiEltCst(11, reference16, reference17)
    constraint8.mode = 0
    reference18 = part1.CreateReferenceFromObject(line2D9)
    reference19 = part1.CreateReferenceFromObject(line2D3)
    reference20 = part1.CreateReferenceFromObject(line2D10)
    constraint9 = constraints1.AddTriEltCst(15, reference18, reference19, reference20)
    constraint9.mode = 0
    reference21 = part1.CreateReferenceFromObject(line2D5)
    reference22 = part1.CreateReferenceFromObject(line2D7)
    reference23 = part1.CreateReferenceFromObject(line2D10)
    constraint10 = constraints1.AddTriEltCst(15, reference21, reference22, reference23)
    constraint10.mode = 0
    reference24 = part1.CreateReferenceFromObject(line2D10)
    reference25 = part1.CreateReferenceFromObject(line2D7)
    constraint11 = constraints1.AddBiEltCst(1, reference24, reference25)
    constraint11.mode = 0
    length1 = constraint11.dimension
    length1.Value = 2.5
    reference26 = part1.CreateReferenceFromObject(line2D9)
    constraint12 = constraints1.AddMonoEltCst(5, reference26)
    constraint12.mode = 0
    length2 = constraint12.dimension
    length2.Value = 1  #
    reference27 = part1.CreateReferenceFromObject(line2D3)
    constraint13 = constraints1.AddMonoEltCst(5, reference27)
    constraint13.mode = 0
    length3 = constraint13.dimension
    length3.Value = 1  #
    reference28 = part1.CreateReferenceFromObject(line2D7)
    constraint14 = constraints1.AddMonoEltCst(5, reference28)
    constraint14.mode = 0
    length4 = constraint14.dimension
    length4.Value = 3
    reference31 = part1.CreateReferenceFromObject(point2D8)
    reference32 = part1.CreateReferenceFromObject(line2D10)
    constraint16 = constraints1.AddBiEltCst(1, reference31, reference32)
    constraint16.mode = 0
    length5 = constraint16.dimension
    length5.Value = 1
    reference29 = part1.CreateReferenceFromObject(line2D10)
    reference30 = part1.CreateReferenceFromObject(line2D11)
    constraint15 = constraints1.AddBiEltCst(6, reference29, reference30)
    constraint15.mode = 0
    constraint15.AngleSector = 0
    angle1 = constraint15.dimension
    angle1.Value = AngleNumber
    ElementReference5 = constraint16
    ElementReference6 = constraint15
    if CircleLineType == False and LineType == 1:
        ElementPoint17 = point2D1
        ElementPoint19 = point2D8
    else:
        ElementPoint17 = point2D8
        ElementPoint19 = point2D1
    ElementSketch.CloseEdition()
    part1.InWorkObject = ElementHybridBody
    part1.Update()
    part1.InWorkObject = ElementSketch
    return ElementReference5, ElementReference6, ElementPoint17, ElementPoint19


def SketchHidePoint(MainSketch, CirclcPoint, X, Y, Construct, ElementDocument, ElementSketch):
    part1 = ElementDocument.Part
    part1.InWorkObject = MainSketch
    factory2D1 = MainSketch.OpenEdition()
    geometricElements1 = MainSketch.GeometricElements
    axis2D1 = geometricElements1.Item("AbsoluteAxis")
    line2D5 = axis2D1.getItem("HDirection")
    line2D6 = axis2D1.getItem("VDirection")
    point2D1 = factory2D1.CreatePoint(0, 0)
    point2D1.ReportName = 1
    if Construct == "True":
        point2D1.Construction = True
    elif Construct == "False":
        point2D1.Construction = False
    ElementPoint29 = point2D1
    ElementPoint30 = CirclcPoint
    (X) = SketchBuildCallout(ElementSketch, "Horizontal", "Binding", X, ElementDocument, ElementPoint29, ElementPoint30)
    (Y) = SketchBuildCallout(ElementSketch, "Vertical", "Binding", Y, ElementDocument, ElementPoint29, ElementPoint30)
    ElementReference30 = point2D1
    MainSketch.CloseEdition()
    part1.InWorkObject = MainSketch
    return ElementReference30


def SketchRectangle(MainSketch, length, wide, ElementDocument, ElementHybridBody):
    part1 = ElementDocument.Part
    part1.InWorkObject = MainSketch
    factory2D1 = MainSketch.OpenEdition()
    geometricElements1 = MainSketch.GeometricElements
    axis2D1 = geometricElements1.Item("AbsoluteAxis")
    axis2D1.ReportName = 1
    line2D5 = axis2D1.getItem("HDirection")
    line2D5.ReportName = 2
    line2D6 = axis2D1.getItem("VDirection")
    line2D6.ReportName = 3
    point2D1 = factory2D1.CreatePoint(-10, 20)
    point2D1.ReportName = 4
    point2D2 = factory2D1.CreatePoint(10, 20)
    point2D2.ReportName = 5
    point2D3 = factory2D1.CreatePoint(10, -20)
    point2D3.ReportName = 6
    point2D4 = factory2D1.CreatePoint(-10, -20)
    point2D4.ReportName = 7
    point2D5 = factory2D1.CreatePoint(0, 0)
    point2D5.Construction = True
    point2D5.ReportName = 8
    line2D1 = factory2D1.CreateLine(-10, 20, 10, 20)
    line2D1.StartPoint = point2D1
    line2D1.EndPoint = point2D2
    line2D2 = factory2D1.CreateLine(10, 20, 10, -20)
    line2D2.EndPoint = point2D2
    line2D2.StartPoint = point2D3
    line2D3 = factory2D1.CreateLine(10, -20, -10, -20)
    line2D3.StartPoint = point2D3
    line2D3.EndPoint = point2D4
    line2D4 = factory2D1.CreateLine(-10, -20, -10, 20)
    line2D4.EndPoint = point2D4
    line2D4.StartPoint = point2D1
    constraints1 = MainSketch.Constraints
    reference1 = part1.CreateReferenceFromObject(line2D1)
    HD_reference = part1.CreateReferenceFromObject(line2D5)
    VD_reference = part1.CreateReferenceFromObject(line2D6)
    base_point_reference = part1.CreateReferenceFromObject(point2D5)
    constraint1 = constraints1.AddBiEltCst(10, reference1, HD_reference)
    constraint1.mode = 0
    reference3 = part1.CreateReferenceFromObject(line2D3)
    constraint2 = constraints1.AddBiEltCst(10, reference3, HD_reference)
    constraint2.mode = 0
    reference5 = part1.CreateReferenceFromObject(line2D2)
    constraint3 = constraints1.AddBiEltCst(13, reference5, VD_reference)
    constraint3.mode = 0
    reference7 = part1.CreateReferenceFromObject(line2D4)
    constraint4 = constraints1.AddBiEltCst(13, reference7, VD_reference)
    constraint4.mode = 0
    constraint5 = constraints1.AddMonoEltCst(5, reference1)
    length1 = constraint5.dimension
    length1.Value = length
    constraint6 = constraints1.AddMonoEltCst(5, reference7)
    length2 = constraint6.dimension
    length2.Value = wide
    constraint7 = constraints1.AddBiEltCst(1, base_point_reference, reference1)
    length3 = constraint7.dimension
    length3.Value = length2.Value / 2
    constraint8 = constraints1.AddBiEltCst(1, base_point_reference, reference7)
    length4 = constraint8.dimension
    length4.Value = length1.Value / 2
    MainSketch.CloseEdition()
    part1.InWorkObject = ElementHybridBody
    part1.Update()
    element_point1 = point2D5
    element_point10 = point2D5
    element_point11 = point2D1
    element_point12 = point2D2
    element_point13 = point2D3
    element_point14 = point2D4
    element_line1 = line2D1
    element_line2 = line2D2
    element_line3 = line2D3
    element_line4 = line2D4

    return element_point1, element_line1, element_line2, element_line3, element_line4


def TranslateElement(MElement, distance, direction, ElementDocument, ElementHybridBody, ElementBody, SketchPosition):
    part1 = ElementDocument.Part
    hybridShapeFactory1 = part1.HybridShapeFactory
    if direction == "X":
        hybridShapeDirection1 = hybridShapeFactory1.AddNewDirectionByCoord(1, 0, 0)
    else:
        pass
    hybridShapeTranslate1 = hybridShapeFactory1.AddNewEmptyTranslate()
    reference1 = part1.CreateReferenceFromObject(MElement)
    hybridShapeTranslate1.ElemToTranslate = reference1
    hybridShapeTranslate1.VectorType = 0
    hybridShapeTranslate1.direction = hybridShapeDirection1
    hybridShapeTranslate1.DistanceValue = distance
    hybridShapeTranslate1.VolumeResult = False
    if SketchPosition == "Body":
        ElementBody.InsertHybridShape(hybridShapeTranslate1)
    elif SketchPosition == "Hyhridbody":
        ElementHybridBody.AppendHybridShape(hybridShapeTranslate1)
    reference3 = part1.CreateReferenceFromObject(hybridShapeTranslate1)
    hybridShapeCurveExplicit2 = hybridShapeFactory1.AddNewCurveDatum(reference3)
    if SketchPosition == "Body":
        ElementBody.InsertHybridShape(hybridShapeCurveExplicit2)
    elif SketchPosition == "Hyhridbody":
        ElementHybridBody.AppendHybridShape(hybridShapeCurveExplicit2)
    ElementReference1 = hybridShapeCurveExplicit2
    part1.InWorkObject = hybridShapeCurveExplicit2
    part1.Update()
    hybridShapeFactory1.DeleteObjectForDatum(reference3)
    return ElementReference1


def HoleSimpleD(d, dimension, direction, ElementDocument, ElementBody, ElementReference11, ElementReference12):
    part1 = ElementDocument.Part
    parameters1 = part1.Parameters
    shapeFactory1 = part1.ShapeFactory
    part1.InWorkObject = ElementBody  # 目前工作位置=草圖
    reference1 = part1.CreateReferenceFromObject(ElementReference11)
    reference2 = part1.CreateReferenceFromObject(ElementReference12)
    hole1 = shapeFactory1.AddNewHoleFromRefPoint(reference1, reference2, 15)
    # ------------------------------------------------------------↓     孔的型態設定
    hole1.Type = 0  # 直孔  (沉頭孔令外再用草圖挖)
    hole1.ThreadingMode = 1  # (螺紋孔catThreadedHoleThreading  OR 無螺紋孔1)
    hole1.ThreadSide = 0  # 左OR右螺紋
    hole1.AnchorMode = 0
    hole1.BottomType = 0
    # ------------------------------------------------------------↑
    # =============================================↓   極限設定
    limit2 = hole1.BottomLimit
    limit2.LimitMode = 0  # 未貫穿(盲孔)
    # =============================================↑
    # =============================================↓   直徑設定
    length1 = hole1.Diameter
    length1.Value = d / 2 - 0.5
    # =============================================↑
    # =============================================↓   牙深設定 孔深>牙深(否則錯誤)
    length2 = limit2.dimension  # 孔深
    length2.Value = dimension
    # =============================================↑
    if direction == 1:
        hole1.Reverse()
    sketch1 = hole1.sketch
    selection1 = ElementDocument.Selection
    # =============隱藏元素===============
    visPropertySet1 = selection1.VisProperties
    selection1.Clear()
    selection1.add(sketch1)
    visPropertySet1 = visPropertySet1.Parent
    visPropertySet1.SetShow(1)
    selection1.Clear()
    # =============隱藏元素===============
    part1.UpdateObject(hole1)
    return hole1


def MarkLineConstraint(MainSketch, SubSketch1, SubSketch2, GEname1, GEname2, PMname, ElementDocument,
                       ElementHybridBody):
    part1 = ElementDocument.part
    part1.InWorkObject = MainSketch
    part1.Update()
    factory2D1 = MainSketch.OpenEdition()
    L1 = SubSketch1.GeometricElements
    L2 = SubSketch2.GeometricElements
    line2D1 = L1.Item(GEname1)
    line2D2 = L2.Item(GEname2)
    reference1 = part1.CreateReferenceFromObject(line2D1)
    geometricElements1 = factory2D1.CreateProjections(reference1)
    geometry2D1 = geometricElements1.Item("Mark.1")
    reference2 = part1.CreateReferenceFromObject(line2D2)
    geometricElements2 = factory2D1.CreateProjections(reference2)
    geometry2D2 = geometricElements2.Item("Mark.1")
    reference3 = part1.CreateReferenceFromObject(geometry2D1)
    reference4 = part1.CreateReferenceFromObject(geometry2D2)
    constraints1 = MainSketch.Constraints
    constraint1 = constraints1.AddBiEltCst(1, reference3, reference4)
    constraint1.mode = 1
    length1 = constraint1.dimension
    # -----------------------建立參數↓
    parameters1 = part1.Parameters
    para1 = parameters1.CreateDimension(PMname, "LENGTH", length1.Value)
    # -----------------------建立參數↑
    relations1 = part1.Relations
    formula1 = relations1.CreateFormula("Formula.1", "", para1, str(length1.Name)[6:len(length1.Name)])
    MainSketch.CloseEdition()
    part1.InWorkObject = ElementHybridBody
    part1.Update()


def MarkLineLengthConstraint(MainSketch, SubSketch1, GEname1, PMname, ElementDocument, ElementHybridBody):
    part1 = ElementDocument.Part
    part1.InWorkObject = MainSketch
    part1.Update()
    factory2D1 = MainSketch.OpenEdition()
    L1 = SubSketch1.GeometricElements
    line2D1 = L1.Item(GEname1)
    reference1 = part1.CreateReferenceFromObject(line2D1)
    geometricElements1 = factory2D1.CreateProjections(reference1)
    geometry2D1 = geometricElements1.Item("Mark.1")
    reference3 = part1.CreateReferenceFromObject(geometry2D1)
    constraints1 = MainSketch.Constraints
    constraint1 = constraints1.AddMonoEltCst(5, reference3)
    constraint1.mode = 1
    length1 = constraint1.dimension
    # -----------------------建立參數↓
    parameters1 = part1.Parameters
    para1 = parameters1.CreateDimension(PMname, "LENGTH", length1.Value)
    # -----------------------建立參數↑
    relations1 = part1.Relations
    formula1 = relations1.CreateFormula("Formula.1", "", para1, str(length1.Name)[6:len(length1.Name)])
    MainSketch.CloseEdition()
    part1.InWorkObject = ElementHybridBody
    part1.Update()


def MarkLineAxisConstraint(MainSketch, SubSketch1, SubSketch2, GEname1, GEname2, PMname, ElementDocument,
                           ElementHybridBody):
    part1 = ElementDocument.Part
    part1.InWorkObject = MainSketch
    part1.Update()
    factory2D1 = MainSketch.OpenEdition()
    L1 = SubSketch1.GeometricElements
    L2 = SubSketch2.GeometricElements
    line2D1 = L1.Item(GEname1)
    axis2D2 = L2.Item("AbsoluteAxis")
    if GEname2 == "HD":
        line2D2 = axis2D2.getItem("VDirection")
    else:
        line2D2 = axis2D2.getItem("HDirection")
    reference1 = part1.CreateReferenceFromObject(line2D1)
    geometricElements1 = factory2D1.CreateProjections(reference1)
    geometry2D1 = geometricElements1.Item("Mark.1")
    reference3 = part1.CreateReferenceFromObject(geometry2D1)
    reference4 = part1.CreateReferenceFromObject(line2D2)
    constraints1 = MainSketch.Constraints
    constraint1 = constraints1.AddBiEltCst(1, reference3, reference4)
    constraint1.mode = 1
    MainSketch.CloseEdition()
    length1 = constraint1.dimension
    # -----------------------建立參數↓
    parameters1 = part1.Parameters
    para1 = parameters1.CreateDimension(PMname, "LENGTH", length1.Value)
    # -----------------------建立參數↑
    relations1 = part1.Relations
    formula1 = relations1.CreateFormula("Formula.1", "", para1, str(length1.Name)[6:len(length1.Name)])
    MainSketch.CloseEdition()
    part1.InWorkObject = ElementHybridBody
    part1.Update()


def MarkLineAxisPointConstraint(MainSketch, SubSketch2, GEname2, PMname, ElementDocument, ElementHybridBody):
    part1 = ElementDocument.Part
    part1.InWorkObject = MainSketch
    part1.Update()
    factory2D1 = MainSketch.OpenEdition()
    hybridShapes1 = ElementHybridBody.HybridShapes
    L2 = SubSketch2.GeometricElements
    axis2D2 = L2.Item("AbsoluteAxis")
    if GEname2 == "HD":
        line2D2 = axis2D2.getItem("VDirection")
    else:
        line2D2 = axis2D2.getItem("HDirection")
    reference1 = hybridShapes1.Item("plate_centor_point")
    geometricElements1 = factory2D1.CreateProjections(reference1)
    geometry2D1 = geometricElements1.Item("Mark.1")
    reference3 = part1.CreateReferenceFromObject(geometry2D1)
    reference4 = part1.CreateReferenceFromObject(line2D2)
    constraints1 = MainSketch.Constraints
    constraint1 = constraints1.AddBiEltCst(1, reference3, reference4)
    constraint1.mode = 1
    MainSketch.CloseEdition()
    length1 = constraint1.dimension
    # -----------------------建立參數↓
    parameters1 = part1.Parameters
    para1 = parameters1.CreateDimension(PMname, "LENGTH", length1.Value)
    # -----------------------建立參數↑
    relations1 = part1.Relations
    formula1 = relations1.CreateFormula("Formula.1", "", para1, str(length1.Name)[6:len(length1.Name)])
    MainSketch.CloseEdition()
    part1.InWorkObject = ElementHybridBody
    part1.Update()


def Project(SketchName, LineName, ElementDocument):
    part1 = ElementDocument.Part
    hybridShapeFactory1 = part1.HybridShapeFactory
    hybridBodies1 = part1.HybridBodies
    hybridBody1 = hybridBodies1.Item("die")
    sketches1 = hybridBody1.HybridSketches
    sketch1 = sketches1.Item(SketchName)
    reference1 = part1.CreateReferenceFromObject(sketch1)
    originElements1 = part1.OriginElements
    hybridShapePlaneExplicit1 = originElements1.PlaneXY
    reference2 = part1.CreateReferenceFromObject(hybridShapePlaneExplicit1)
    hybridShapeProject1 = hybridShapeFactory1.AddNewProject(reference1, reference2)
    hybridShapeProject1.SolutionType = 0
    hybridShapeProject1.Normal = True
    hybridShapeProject1.SmoothingType = 0
    hybridBody1.AppendHybridShape(hybridShapeProject1)
    part1.InWorkObject = hybridShapeProject1
    part1.Update()
    reference3 = part1.CreateReferenceFromObject(hybridShapeProject1)
    hybridShapeCurveExplicit1 = hybridShapeFactory1.AddNewCurveDatum(reference3)
    hybridBody1.AppendHybridShape(hybridShapeCurveExplicit1)
    hybridShapeCurveExplicit1.Name = LineName
    part1.Update()
    hybridShapeFactory1.DeleteObjectForDatum(reference3)


def Del1(SketchName, Status, final, ElementDocument):
    part1 = ElementDocument.Part
    hybridShapeFactory1 = part1.HybridShapeFactory
    hybridBodies1 = part1.HybridBodies
    hybridBody1 = hybridBodies1.Item("die")
    sketches1 = hybridBody1.HybridSketches
    selection1 = ElementDocument.Selection
    try:
        sketch1 = sketches1.Item(SketchName)
        selection1.Add(sketch1)
        if Status == False:
            selection1.Delete()
            final = True
    except:
        if Status == False:
            selection1.Delete()
            final = True
    return final


def OriginalPoint(element, environment_element, ElementDocument, ElementBody, ElementHybridBody):
    part1 = ElementDocument.Part
    bodies1 = part1.Bodies
    sketches1 = ElementBody.Sketches
    hybridShapes1 = ElementHybridBody.HybridShapes
    hybridShapes2 = ElementBody.HybridShapes
    hybridShapeFactory1 = part1.HybridShapeFactory
    # -----------------------------------------------------------------建立向量參數
    X_min = hybridShapeFactory1.AddNewDirectionByCoord(1, 0, 0)  # X方向  <- = +  -> = -     (X,Y,Z)
    X_max = hybridShapeFactory1.AddNewDirectionByCoord(-1, 0, 0)  # Y方向  ↓ = +  ↑ = -     (X,Y,Z)
    Y_min = hybridShapeFactory1.AddNewDirectionByCoord(0, 1, 0)  # Y方向  ↓ = +  ↑ = -     (X,Y,Z)
    Y_max = hybridShapeFactory1.AddNewDirectionByCoord(0, -1, 0)  # Y方向  ↓ = +  ↑ = -     (X,Y,Z)
    zero = hybridShapeFactory1.AddNewDirectionByCoord(0, 0, 0)  # 無方向     (X,Y,Z)
    # ------------------------------------------------------------------建立向量參數
    reference2 = part1.CreateReferenceFromObject(element)
    # ------------------------------------------------------------------建立極值點1
    hybridShapeExtremum1 = hybridShapeFactory1.AddNewExtremum(reference2, X_min, 0)
    hybridShapeExtremum1.Direction2 = Y_min
    hybridShapeExtremum1.ExtremumType2 = 0
    ElementHybridBody.AppendHybridShape(hybridShapeExtremum1)
    hybridShapeExtremum1.Name = str(environment_element) + "_basis_point"
    ElementReference1 = hybridShapeExtremum1
    return ElementReference1


def ChangeExtremum(reference_Boundary, c):
    catapp = win32.Dispatch("CATIA.Application")
    partDocument1 = catapp.ActiveDocument
    part1 = partDocument1.Part
    hybridShapeFactory1 = part1.HybridShapeFactory
    # -------------------------------------------------------------------建立向量參數
    X_min = hybridShapeFactory1.AddNewDirectionByCoord(1, 0, 0)  # X方向  <- = +  -> = -     (X,Y,Z)
    X_max = hybridShapeFactory1.AddNewDirectionByCoord(-1, 0, 0)  # Y方向  ↓ = +  ↑ = -     (X,Y,Z)
    Y_min = hybridShapeFactory1.AddNewDirectionByCoord(0, 1, 0)  # Y方向  ↓ = +  ↑ = -     (X,Y,Z)
    Y_max = hybridShapeFactory1.AddNewDirectionByCoord(0, -1, 0)  # Y方向  ↓ = +  ↑ = -     (X,Y,Z)
    zero = hybridShapeFactory1.AddNewDirectionByCoord(0, 0, 0)  # 無方向     (X,Y,Z)
    # -------------------------------------------------------------------建立向量參數
    hybridBodies1 = part1.HybridBodies
    hybridBody1 = hybridBodies1.Item("die")
    hybridShapes1 = hybridBody1.HybridShapes
    hybridShapeExtremum1 = hybridShapes1.Item("measure_point")
    reference1 = part1.CreateReferenceFromObject(reference_Boundary)
    hybridShapeExtremum1.ReferenceElement = reference1
    if c == 1:
        hybridShapeExtremum1.direction = X_min
        hybridShapeExtremum1.direction = X_min
        hybridShapeExtremum1.Direction2 = Y_min
    elif c == 2:
        hybridShapeExtremum1.direction = X_max
        hybridShapeExtremum1.Direction2 = Y_max
    else:
        raise ValueError('Incorrect Target Type')
    part1.Update()


def ExcelSearch(ExcelName, excel_Sheet_name, Row_string_serch, Column_string_serch):
    DieRuleRoot = str(gvar.die_rule_path + ExcelName + '.xlsx')
    workbook = openpyxl.load_workbook(DieRuleRoot)
    sheet = workbook[excel_Sheet_name]
    x = 0
    y = 0
    for r in range(1, sheet.max_row + 1):
        x = x + 1
        v = sheet.cell(row=r, column=1).value
        if v == Row_string_serch:
            break
    for c in range(1, sheet.max_column + 1):
        y = y + 1
        v = sheet.cell(row=1, column=c).value
        if v == Column_string_serch:
            break
    SearchResult = sheet.cell(row=x, column=y).value
    return SearchResult


def FormulaChange(body_name2, formula_name1, line_name1, sketch_name1):
    catapp = win32.Dispatch("CATIA.Application")
    document1 = catapp.ActiveDocument
    partDocument1 = catapp.ActiveDocument
    part1 = partDocument1.Part
    parameters1 = part1.Parameters
    bodies1 = part1.Bodies
    body1 = bodies1.Item(body_name2)
    cut_cavity_machining_explanation_shape = int()
    # '------------------------------------------------------------↑
    cut_cavity_machining_explanation_shape = cut_cavity_machining_explanation_shape + 1
    # '--------------------------------------------------加工說明
    # '------------------------------------------------------------↓置換草圖
    parameters1.Item(formula_name1).OptionalRelation.Modify("die\\" + line_name1)  # 草圖置換
    # '------------------------------------------------------------↑
    # '------------------------------------------------------------↓宣告草圖
    sketches1 = body1.Sketches
    sketch1 = sketches1.Item(sketch_name1)
    part1.UpdateObject(sketch1)


def F_projection(plane, sketch, line, body_name1, body_name2):
    catapp = win32.Dispatch("CATIA.Application")
    document1 = catapp.ActiveDocument
    partDocument1 = catapp.ActiveDocument
    part1 = partDocument1.Part
    hybridShapeFactory1 = part1.HybridShapeFactory
    bodies1 = part1.Bodies
    body1 = bodies1.Item(body_name1)
    body2 = bodies1.Item(body_name2)
    part1.InWorkObject = body1
    sketches1 = body2.Sketches
    sketch1 = sketches1.Item(sketch)
    reference1 = part1.CreateReferenceFromObject(sketch1)
    hybridShapes1 = body1.HybridShapes
    hybridShapePlaneOffset1 = hybridShapes1.Item(plane)
    reference2 = part1.CreateReferenceFromObject(hybridShapePlaneOffset1)
    hybridShapeProject1 = hybridShapeFactory1.AddNewProject(reference1, reference2)
    hybridShapeProject1.SolutionType = 0
    hybridShapeProject1.Normal = True
    hybridShapeProject1.SmoothingType = 0
    body1.InsertHybridShape(hybridShapeProject1)
    part1.InWorkObject = hybridShapeProject1
    part1.Update()
    reference3 = part1.CreateReferenceFromObject(hybridShapeProject1)
    hybridShapeCurveExplicit1 = hybridShapeFactory1.AddNewCurveDatum(reference3)
    body1.InsertHybridShape(hybridShapeCurveExplicit1)
    part1.InWorkObject = hybridShapeCurveExplicit1
    part1.InWorkObject.Name = line
    part1.Update()
    hybridShapeFactory1.DeleteObjectForDatum(reference3)
    selection1 = partDocument1.Selection
    selection1.Clear()
    selection1.Add(hybridShapeProject1)
    selection1.VisProperties.SetShow(1)  # 1為隱藏,0為顯示
    selection1.Clear()


def F_Excavation(plane, line, body_name1):
    catapp = win32.Dispatch("CATIA.Application")
    document1 = catapp.ActiveDocument
    partDocument1 = catapp.ActiveDocument
    part1 = partDocument1.Part
    parameters1 = part1.Parameters
    shapeFactory1 = part1.ShapeFactory
    hybridShapeFactory1 = part1.HybridShapeFactory
    bodies1 = part1.Bodies
    body1 = bodies1.Item(body_name1)
    part1.InWorkObject = body1
    hybridShapes1 = body1.HybridShapes
    hybridShapeCurveExplicit1 = parameters1.Item(line)
    # '=============================================↓   建立條件
    reference1 = hybridShapes1.Item(plane)  # '指定(平面)為條件
    # '=============================================↑
    pocket1 = shapeFactory1.AddNewPocketFromRef(reference1, 20)
    reference2 = part1.CreateReferenceFromObject(hybridShapeCurveExplicit1)
    pocket1.SetProfileElement(reference2)  # '挖除形狀元素    草圖1
    # '=============================================↓   極限設定
    limit1 = pocket1.FirstLimit
    limit1.LimitMode = 2  # '完全挖除
    # ==============================================↑   極限設定


def FUN_pad_gap(gap_file_name, EX_file, sketch_neme, deep, gap, body_name2, body_name3, parameter_name5, SketchPosition,
                now_data_number):  # 貼上元素   body_name(2)=copy"body1"  body_name(3)=paste "body2" parameter_name(5)="bending_cavity_parameter" deep="max"=>挖到底
    length = [] * 20
    catapp = win32.Dispatch("CATIA.Application")
    documents1 = catapp.Documents
    partDocument1 = documents1.Open(gvar.save_path + gap_file_name + ".CATPart")
    (ElementProduct, ElementDocument, ElementBody, ElementHybridBody) = environment_set(gap_file_name, body_name2,
                                                                                        "die")  # 環境設定
    part1 = ElementDocument.Part
    sketches1 = ElementBody.Sketches
    sketch1 = sketches1.Item(sketch_neme)
    hybridShape1 = ElementHybridBody.HybridShapes.Item("down_plane")  # 宣告平面
    parametersParent = part1.Parameters
    parameters1 = parametersParent.RootParameterSet.ParameterSets.getItem(parameter_name5)
    parameter5 = parameters1.DirectParameters.Item("gap")
    length[1] = parameter5
    length[1].Value = gap
    element_Reference11 = sketch1
    element_Reference12 = hybridShape1
    (ElementLine5) = ProjectionLine(element_Reference11, element_Reference12, ElementDocument, ElementHybridBody,
                                    ElementBody, SketchPosition,
                                    'True')  # 投影線段  element_Reference(11)=投影之元素 #element_Reference(12)=plane  element_line(5) 為out
    ElementLine5.Name = sketch_neme + "_line" + now_data_number
    selection1 = ElementDocument.Selection
    selection1.Clear()
    selection1.Add(ElementLine5)
    selection1.Copy()
    if EX_file == "Data1":
        partDocument2 = documents1.Item(EX_file + ".CATPart")
    else:
        partDocument2 = documents1.Open(gvar.save_path + EX_file + ".CATPart")
    (ElementProduct, ElementDocument, ElementBody, ElementHybridBody) = environment_set(EX_file, body_name3,
                                                                                        "die")  # 環境設定
    part1 = ElementDocument.Part
    selection1 = ElementDocument.Selection
    selection1.Clear()
    selection1.Add(ElementBody)
    selection1.Paste()
    partDocument1.Close()
    hybridShape2 = ElementBody.HybridShapes.Item(sketch_neme & "_line" + str(now_data_number))  # 宣告投影線
    shapeFactory1 = part1.ShapeFactory
    reference1 = part1.CreateReferenceFromObject(hybridShape2)
    part1.InWorkObject = ElementBody
    pocket1 = shapeFactory1.AddNewPocketFromRef(reference1, 20)
    element_Reference20 = pocket1
    if deep == "max":
        element_Reference20.FirstLimit.LimitMode = 2
    else:
        element_Reference20.FirstLimit.dimension = deep
    part1.Update()
    if EX_file != "Data1":
        partDocument2.save()
        partDocument2.Close()


def material_type_palte_sketch(E_plane, M_plate_length, M_plate_wide, center_X, center_Y, ElementDocument, ElementBody,
                               ElementHybridbody, SketchPosition):
    hybridShape1 = ElementHybridbody.HybridShapes.Item("upper_die_seat_line")
    (ElementReference1) = ExtremumPoint("X_min", "Y_min", "Z_max", 2, hybridShape1, ElementDocument, ElementBody,
                                        ElementHybridbody)
    ElementReference2 = ElementReference1
    ElementReference2.Name = "seat_plate_min"
    (ElementReference1) = ExtremumPoint("X_max", "Y_max", "Z_max", 2, hybridShape1, ElementDocument, ElementBody,
                                        ElementHybridbody)
    ElementReference3 = ElementReference1
    ElementReference3.Name = "seat_plate_min"
    (ElementSketch) = BuildSketch("plate_size", E_plane, ElementDocument, SketchPosition, ElementBody,
                                  ElementHybridbody)
    (ElementPoint1,element_line1, element_line2, element_line3, element_line4) = SketchRectangle(ElementSketch, M_plate_length, M_plate_wide, ElementDocument,
                                      ElementHybridbody)
    # =============隱藏元素===============
    selection1 = ElementDocument.Selection
    visPropertySet1 = selection1.VisProperties
    selection1.Clear()
    selection1.add(ElementSketch)
    visPropertySet1 = visPropertySet1.Parent
    visPropertySet1.SetShow(1)
    selection1.Clear()
    # =============隱藏元素===============
    ElementPoint2 = ElementReference2
    ElementPoint3 = ElementReference2
    ElementPoint4 = ElementReference3
    seat_length = 0
    seat_wide = 0
    (seat_length) = SketchBuildCallout(ElementSketch, "Horizontal", "Callout", seat_length, ElementDocument,
                                       ElementPoint3, ElementPoint4)
    (seat_wide) = SketchBuildCallout(ElementSketch, "Vertical", "Callout", seat_wide, ElementDocument,
                                     ElementPoint3, ElementPoint4)
    SketchBuildCallout(ElementSketch, "Horizontal", "Binding", center_X, ElementDocument, ElementPoint1,
                       ElementPoint2)
    SketchBuildCallout(ElementSketch, "Vertical", "Binding", center_Y, ElementDocument, ElementPoint1,
                       ElementPoint2)
    return ElementSketch


def material_tpye_palte_sketch(E_plane, M_plate_length, M_plate_wide, center_X, center_Y, ElementDocument, ElementBody,
                               ElementHybridbody):  # (平面的宣告,模板長,模板寬,中心座標_X,中心座標_Y)
    Sketch_position = "A"
    hybridShape1 = ElementHybridbody.HybridShapes.Item("upper_die_seat_line")  # 宣告元素(當下模板的取線)
    (ElementReference1) = ExtremumPoint("X_min", "Y_min", "Z_max", 2, hybridShape1, ElementDocument, ElementBody,
                                        ElementHybridbody)  # 建立極值點   (極值方向1,極值方向2,極值方向3,需要幾個方向,在哪個元素上的極值)  element_Reference(1)為OUT
    ElementReference2 = ElementReference1
    ElementReference2.Name = "seat_plate_min"  # 建立最小點
    (ElementReference1) = ExtremumPoint("X_max", "Y_max", "Z_max", 2, hybridShape1, ElementDocument, ElementBody,
                                        ElementHybridbody)  # 建立極值點   (極值方向1,極值方向2,極值方向3,需要幾個方向,在哪個元素上的極值)  element_Reference(1)為OUT
    ElementReference3 = ElementReference1
    ElementReference3.Name = "seat_plate_max"  # 建立最大點
    (ElementSketch) = BuildSketch("plate_size", E_plane, ElementDocument, Sketch_position, ElementBody,
                                  ElementHybridbody)  # (sketch名稱,產生sketch的平面)  element_sketch(1)為產生出來的草圖
    (ElementPoint1,element_line1, element_line2, element_line3, element_line4) = SketchRectangle(ElementSketch, M_plate_length, M_plate_wide, ElementDocument,
                                      ElementHybridbody)  # (草圖陳述句,長,寬)  element_point(1)=>output中心點之陳述句  [需經環境副程式跑過]
    # =============隱藏元素===============
    part1 = ElementDocument.Part
    selection1 = part1.Selection
    visPropertySet1 = selection1.VisProperties
    selection1.Clear()
    selection1.add(ElementSketch)
    visPropertySet1 = visPropertySet1.Parent
    visPropertySet1.SetShow(1)
    selection1.Clear()
    # =============隱藏元素===============
    ElementPoint2 = ElementReference2
    ElementPoint3 = ElementReference2
    ElementPoint4 = ElementReference2
    seat_length = float()
    seat_wide = float()
    (seat_length) = SketchBuildCallout(ElementSketch, "Horizontal", "Callout", seat_length, ElementDocument,
                                       ElementPoint3, ElementPoint4)  # 建立標註(草圖陳述句,標註方向,式標OR拘束,改變OR讀取之數值,標註兩點)
    (seat_wide) = SketchBuildCallout(ElementSketch, "Vertical", "Callout", seat_wide, ElementDocument, ElementPoint3,
                                     ElementPoint4)  # 建立標註(草圖陳述句,標註方向,式標OR拘束,改變OR讀取之數值,標註兩點)#element_point(now_point)[點之陳述句]和element_point(now_point+1進行標註)direction="Horizontal"or"Vertical"or"free"
    SketchBuildCallout(ElementSketch, "Horizontal", "Binding", center_X, ElementDocument, ElementPoint1,
                       ElementPoint2)  # 建立標註(草圖陳述句,標註方向,式標OR拘束,改變OR讀取之數值,標註兩點)
    SketchBuildCallout(ElementSketch, "Vertical", "Binding", center_Y, ElementDocument, ElementPoint1,
                       ElementPoint2)  # 建立標註(草圖陳述句,標註方向,式標OR拘束,改變OR讀取之數值,標註兩點)


def hole_simple_M(M, dimension, direction, ElementDocument, ElementBody, ElementReference11, ElementReference12):
    part1 = ElementDocument.Part
    parameters1 = part1.Parameters
    shapeFactory1 = part1.ShapeFactory
    part1.InWorkObject = ElementBody  # 目前工作位置=草圖
    reference1 = part1.CreateReferenceFromObject(ElementReference11)
    reference2 = part1.CreateReferenceFromObject(ElementReference12)
    hole1 = shapeFactory1.AddNewHoleFromRefPoint(reference1, reference2, 15)
    # ------------------------------------------------------------↓     孔的型態設定
    hole1.Type = 0  # (catSimpleHole)  # 直孔  (沉頭孔令外再用草圖挖)
    hole1.ThreadingMode = 0  # (catThreadedHoleThreading)  # (螺紋孔0  , 無螺紋孔1)
    hole1.ThreadSide = 0  # (catRightThreadSide)  # 左OR右螺紋
    hole1.AnchorMode = 0  # (catExtremPointHoleAnchor)
    hole1.BottomType = 0  # (catFlatHoleBottom)
    hole1.CreateStandardThreadDesignTable(1)  # (catHoleMetricThickPitch)
    # ------------------------------------------------------------↑
    strParam1 = hole1.HoleThreadDescription
    strParam1.Value = M
    # =============================================↓   極限設定
    limit2 = hole1.BottomLimit
    limit2.LimitMode = 0  # (catOffsetLimit)  # 未貫穿(盲孔)
    # =============================================↓   牙深設定 孔深>牙深(否則錯誤)
    length2 = limit2.dimension  # 孔深
    length2.Value = dimension
    # =============================================↓   牙深設定 孔深>牙深(否則錯誤)
    length3 = hole1.ThreadDepth  # 牙深
    length3.Value = length2.Value - 2
    # =============================================↑
    if direction == 1:
        hole1.Reverse()
    sketch1 = hole1.sketch
    selection1 = ElementDocument.Selection
    visPropertySet1 = selection1.VisProperties
    selection1.Clear()
    selection1.Add(sketch1)
    visPropertySet1 = visPropertySet1.Parent
    visPropertySet1.SetShow(1)
    selection1.Clear()
    part1.UpdateObject(hole1)
    part1.Update()
    return hole1
